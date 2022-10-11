# -*- coding: utf-8 -*-
"""
Created on Fri Jul 22 16:55:22 2022

@author: Alireza
"""
import os
import pandas as pd
import numpy as np
import glob

nsda = 9
no_prosumers = 10
header = [i for i in range(1, nsda+1)]
horizon = [t for t in range(16,40)]

MVA = 30
PU_DA = 1/(1000*MVA)


from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
# Electric Vehicles Information
EVs_sheets =["max_soc", "min_soc", "Arrival Times", "Departure Times", "Charging Power", "Arrival SOC", "Charging Efficiency" , "Discharging Efficiency"]

# Prosumers EVs Information
prosumers_Evs = ["EV_soc_up","EV_soc_low", "Arrival", "Depart", "EV_Power", "EV_soc_arr"]

# prosumers TCL Information
prosumers_tcl=['TCL_R', "TCL_COP", "TCL_MAX", "TCL_Beta", "TCL_temp_low", "TCL_temp_up", "Arrival", "Depart"]
TCL_sheets = ['R','COP','Max Consumption', 'Beta', 'Low Temps','Up Temps' , 'Start Times','End Times',]
outside_temp=[27.694803,26.834803,26.594803,25.664803,22.594803,21.394802,20.164803,19.584803,20.334803,16.784803,16.094803,15.764802,14.774801,14.834802,14.184802,14.144801,15.314801,16.694803,19.734802,24.414803,25.384802,26.744802,27.144802,27.524803]


#prosumers Shiftable loads
prosumers_sl =["SL_loads1", "SL_loads2", "SL_low", "SL_up"]
SL_sheets   = ['SL Consumption', 'SL Start', 'SL End']




def create_EVs_input_Data(fileName, fileAdd):
    exclude = ["Charging Efficiency" , "Discharging Efficiency"]
    ch_efficeincy = 0.95
    
    prosumers_file_add = os.path.join("..",  "prosumers_data")
    pr_file_name = "prosumers_profiles_scen_"
    
    # results will save to this file
    evs_excel_add = os.path.join(fileAdd, fileName)
    
    
    index = 0
    for sheet in EVs_sheets:
        dic_data = dict()
        for i in range(1, nsda+1):
            if sheet not in exclude:
                #print(sheet)
                add_file = os.path.join(os.pardir,  "prosumers_data", pr_file_name+str(i)+".csv")
                Evs_Info = pd.read_csv(add_file, usecols=prosumers_Evs, nrows=no_prosumers)
                for info in ["EV_soc_up","EV_soc_low", "EV_Power", "EV_soc_arr"]:
                    Evs_Info[info] = Evs_Info[info]*PU_DA
                dic_data[i] = Evs_Info[prosumers_Evs[index]].tolist()
                
            else:
                eff = [ch_efficeincy for x in range(no_prosumers)]
                dic_data[i] = eff
        
        index += 1
        df= pd.DataFrame().from_dict(dic_data)
        
        
        if os.path.exists(evs_excel_add):
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
                df.to_excel(writer, sheet_name=sheet )
        else:
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl')  as writer: 
                df.to_excel(writer, sheet_name=sheet )
     
        #df.to_excel(evs_excel_add, sheet_name=sheet)
    print("EVs Excell file is saved in: ", evs_excel_add)


def create_inflexible_loads(fileName, fileAdd):
    prosumers_file_add = os.path.join("..",  "prosumers_data")
    pr_file_name = "inflexible_profiles_scen_"
    
    # results will save to this file
    evs_excel_add = os.path.join(fileAdd, fileName)
    
    time_header = ["t="+str(t) for t in horizon ]
    
    data_dic=dict()
    
    for i in range(1, nsda+1):
        add_file = os.path.join(os.pardir,  "prosumers_data", pr_file_name+str(i)+".csv")
        inf_info = pd.read_csv(add_file,  nrows=no_prosumers)
        inf_info = inf_info.sum(axis=0)
        inf_info = inf_info #* PU_DA
        #print(inf_info)
        data_dic[i] = inf_info.tolist()
        
    df = pd.DataFrame().from_dict(data_dic).T
    col_names=df.columns.tolist()
    
    dic_name=dict()
    for i in range(len(col_names)):
        dic_name[col_names[i]] = time_header[i]
        
    df.rename(columns=dic_name, inplace=True)
    
    dic_name=dict()
    for i in range(len(df.index)):
        dic_name[df.index[i]] = 'SDA ' + str(i+1)
        
    df.rename(index=dic_name, inplace=True)
    
    df = (df*PU_DA)/1000
    if os.path.exists(evs_excel_add):
        with pd.ExcelWriter(evs_excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
            df.to_excel(writer, sheet_name='Sheet1' )
    else:
        with pd.ExcelWriter(evs_excel_add, engine='openpyxl')  as writer: 
            df.to_excel(writer, sheet_name="Sheet1" )
    
    print("Iflexible load data as Excell file is saved in: ", evs_excel_add)

def create_TCL_Loads_sheets(fileName, fileAdd):
    prosumers_file_add = os.path.join("..",  "prosumers_data")
    pr_file_name = "prosumers_profiles_scen_"
    
    exclude = ['Outside Temperature']
    
    # results will save to this file
    evs_excel_add = os.path.join(fileAdd, fileName)
    
    index = 0
    for sheet in TCL_sheets:
        dic_data = dict()
        for i in range(1, nsda+1):
            #print(sheet)
            add_file = os.path.join(os.pardir,  "prosumers_data", pr_file_name+str(i)+".csv")
            TCLs_Info = pd.read_csv(add_file, usecols=prosumers_tcl, nrows=no_prosumers)
            dic_data[i] = TCLs_Info[prosumers_tcl[index]].tolist()
                
        
        index += 1
        df= pd.DataFrame().from_dict(dic_data)
        # if(sheet == 'Max Consumption'):
        #     df = df
        
        if os.path.exists(evs_excel_add):
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
                df.to_excel(writer, sheet_name=sheet )
        else:
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl')  as writer: 
                df.to_excel(writer, sheet_name=sheet )
    
    # Creating sheet for outside temprature
    temp_time = ['t='+str(x) for x in range(16,40)]
    index_temp = "θ(°C)"
    dic_data=dict()
    for t in range(len(temp_time)):
        dic_data[temp_time[t]] = [outside_temp[t]]
    
    df = pd.DataFrame().from_dict(dic_data)
    df.rename(index={0:index_temp}, inplace=True)
    
    if os.path.exists(evs_excel_add):
        with pd.ExcelWriter(evs_excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
            df.to_excel(writer, sheet_name="Outside Temperature" )
    else:
        with pd.ExcelWriter(evs_excel_add, engine='openpyxl')  as writer: 
            df.to_excel(writer, sheet_name="Outside Temperature" )
    
    print("_-"*10)
    print("TCL data as Excell file is saved in: ", evs_excel_add)

prosumers_sl =["SL_loads1", "SL_loads2", "SL_low", "SL_up"]
SL_sheets   = ['SL Consumption', 'SL Start', 'SL End']

def create_SL_Loads_sheets(fileName, fileAdd):
    exclude = ["Charging Efficiency" , "Discharging Efficiency"]
    
    prosumers_file_add = os.path.join("..",  "prosumers_data")
    pr_file_name = "prosumers_profiles_scen_"
    
    # results will save to this file
    evs_excel_add = os.path.join(fileAdd, fileName)
    
    index = 0
    
    df = pd.DataFrame()
    
    for sheet in SL_sheets:
        dic_data = dict()
        for i in range(1, nsda+1):
            add_file = os.path.join(os.pardir,  "prosumers_data", pr_file_name+str(i)+".csv")
            SLs_Info = pd.read_csv(add_file, usecols=prosumers_sl, nrows=no_prosumers)
            if sheet == 'SL Consumption':
                temp_df=None
                temp_time = ['SDA_'+str(i)+' t='+str(x) for x in range(16,40)]
                dump_loads = np.ones([no_prosumers, 24]) * -1
                
                cycle = SLs_Info['SL_up'] - SLs_Info['SL_low']
                loads = (SLs_Info['SL_loads1'] + SLs_Info['SL_loads2'])/100
                loads = (loads * PU_DA)/100
                for j in range(no_prosumers):
                    for t in range(SLs_Info['SL_low'][j], SLs_Info['SL_up'][j]+1):
                        dump_loads[j,t-16] = loads[j]
                temp_df=pd.DataFrame(dump_loads, columns = temp_time)
                
            if sheet == 'SL Start':
                dic_data[i] = SLs_Info['SL_low'].tolist()
            if sheet == 'SL End':
                dic_data[i] = SLs_Info['SL_up'].tolist()
        
            if sheet == 'SL Consumption' and df.size > 0:
                df = pd.concat([df, temp_df] , axis=1)
            elif sheet == 'SL Consumption' and df.size ==0:
                df = temp_df.copy(deep=True)
       
        if sheet != 'SL Consumption':
            df= pd.DataFrame().from_dict(dic_data)
        
        if os.path.exists(evs_excel_add):
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
                df.to_excel(writer, sheet_name=sheet )
        else:
            with pd.ExcelWriter(evs_excel_add, engine='openpyxl')  as writer: 
                df.to_excel(writer, sheet_name=sheet )

def get_file_date_extension():
    file_name ="Model_data_DA_1*"
    data_dir = '../Model_CSV/*'
    path = os.getcwd()
    path = os.path.abspath(os.path.join(path, os.pardir))
    path = os.path.join(path, 'Model_CSV')
    file_path = glob.glob(path+"\\"+file_name)[0]
    
    onlyfiles = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
    file = os.path.basename(file_path)
    start = len(file_name)
    end = len(file)-4
    
    return file[start:end]
    
def create_bids_offer(no_das, save_path):
    file_name ="Model_data_DA_"
    excel_file="Strategic DA Quantity Offers.xlsx"
    
    path = os.getcwd()
    path = os.path.abspath(os.path.join(path, os.pardir))
    path = os.path.join(path, 'Model_CSV')
    
    index=["SDA "+str(x) for x in range(1, no_das+1)]
    
    time_col = ["t="+str(x) for x in range(16,40)]
    time_index=dict()
    for i in range(len(time_col)):
        time_index[i] = time_col[i]
    
    posterior = get_file_date_extension()
    
    column_offer=[]
    column_offer.append('DAs_supply_offer\no_t')
    for col in ["CDA"+str(x)+"_supply" for x in range(1, no_das)]:
        column_offer.append(col)
        
    column_bid=[]
    column_bid.append('DAs_demand_bid\nb_t')
    for col in ["CDA"+str(x)+"_demand" for x in range(1, no_das)]:
        column_bid.append(col)
        
    
    data_offer = dict()
    data_bid   = dict()
    
    f_name = file_name + str(1) +"_" + posterior +".csv"
    f_name=os.path.join(path, f_name)
    df = pd.read_csv(f_name, usecols = column_offer + column_bid + ['Iteration'])
    
    for j in range(len(column_offer)):
        data_offer[index[j]]= df[df.Iteration==2][column_offer[j]].tolist()
        data_bid[index[j]] = df[df.Iteration==2][column_bid[j]].tolist()   
     
    df_offer = pd.DataFrame().from_dict(data_offer)
    df_offer.rename(index=time_index, inplace=True)
    df_offer = df_offer.T
    
    df_bid  = pd.DataFrame().from_dict(data_bid)
    df_bid.rename(index=time_index, inplace=True)
    df_bid = df_bid.T
    
    excel_add = os.path.join(save_path, excel_file)
    
    if os.path.exists(excel_add):
        with pd.ExcelWriter(excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
            df_offer.to_excel(writer, sheet_name="Offers" )
    else:
        with pd.ExcelWriter(excel_add, engine='openpyxl')  as writer: 
            df_offer.to_excel(writer, sheet_name="Offers" )
    
    if os.path.exists(excel_add):
        with pd.ExcelWriter(excel_add, engine='openpyxl', mode='a',if_sheet_exists="replace")  as writer: 
            df_bid.to_excel(writer, sheet_name="Bids" )
    else:
        with pd.ExcelWriter(excel_add, engine='openpyxl')  as writer: 
            df_bid.to_excel(writer, sheet_name="Bids" )
    
    # for i in range(1,no_das+1):
    #     f_name = file_name + str(i) +"_" + posterior +".csv"
    #     f_name=os.path.join(path, f_name)
    #     df bid= pd.read_csv(f_name, columns = column_offer + column_bid)
    #     for j in len(column_offer):
    #         data_offer[index[j]]= df[column_offer].tolist()
    #         data_bid[]

save_path = 'Ali Data'
save_path = 'Test_02'
create_EVs_input_Data("EVs.xlsx", save_path)
create_inflexible_loads("Inflexible Consumption.xlsx", save_path)
create_TCL_Loads_sheets("TCL.xlsx", save_path)
create_SL_Loads_sheets("SL.xlsx", save_path)
create_bids_offer(9, save_path)
